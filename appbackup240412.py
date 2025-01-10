import streamlit as st
import pandas as pd
import yfinance as yf
import numpy as np
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
import datetime
from datetime import datetime, timedelta, time
import math
from scipy import stats
from scipy.stats import norm
from scipy.optimize import newton
from openpyxl import Workbook, load_workbook
import requests
import zipfile
import io
import fundamentus
import riskfolio as rp
from statsmodels.regression.linear_model import OLS
from statsmodels.tsa.stattools import adfuller
import statsmodels.api as sm

st.title("Harpa Quant")
st.markdown("""##### Ferramentas quantitativas para o investidor prospectivo.""")
st.markdown("""Escolha à esquerda a ferramenta (no celular, setinha bem em cima à esquerda).""")

col1, col2, col3, col4, col5 = st.columns(5)

# Adicionar conteúdo em cada coluna
with col1:
    st.markdown("[![Twitter](https://img.shields.io/badge/Twitter-%231DA1F2.svg?style=for-the-badge&logo=Twitter&logoColor=white)](https://twitter.com/harpaquant)")

with col2:
    st.markdown("[![Instagram](https://img.shields.io/badge/Instagram-%23E4405F.svg?style=for-the-badge&logo=Instagram&logoColor=white)](https://www.instagram.com/harpaquant)")

with col3:
    st.markdown("")

with col4:
    st.markdown("")

with col5:
    st.markdown("")

st.markdown('---')

st.sidebar.markdown("""
    Acesse nossa [Comunidade no Discord](https://discord.gg/MaF7wZDQvZ) para participar de 
                    discussões no tema. Contato: harpaquant@gmail.com
    """)

st.sidebar.markdown('---')

selected_calculator = st.sidebar.selectbox(
    "Selecione a ferramenta:",
    ("Long Short - Cointegração", "Long Short - Teste seu Par", "PCR - Put Call Ratio", "BDR - Spreads", "Carteiras", "Seguro da Carteira", "Cones de Volatilidade", "Monitor de 5 Dias", "Calculadora de Gregas de Opções", "Calculadoras Black-Scholes-Merton", "Top 10 Fundos Quantitativos")
)

st.sidebar.markdown('---')
st.sidebar.subheader('Ferramentas disponíveis')
st.sidebar.write('Long Short - Cointegração')
st.sidebar.write('Long Short - Teste seu Par')
st.sidebar.write('PCR - Put Call Ratio')
st.sidebar.write('BDR - Spreads')
st.sidebar.write('Carteiras \n\n- Magic Formula de Joel Greenblatt \n\n- Risk Parity' )
st.sidebar.write('Seguro da Carteira')
st.sidebar.write('Cones de Volatilidade')
st.sidebar.write('Monitor de 5 Dias')
st.sidebar.write('Calculadora de Gregas de Opções \n\n- Delta, Gamma, Vega, Theta, Rho ')
st.sidebar.write('Calculadoras Black-Scholes-Merton \n\n- Preço da opção\n\n- Volatilidade implícita')
st.sidebar.write('Top 10 Fundos Quantitativos')

###########################
### BLACK-SCHOLES

if selected_calculator == "Calculadoras Black-Scholes-Merton":
    
    ### Calculadora Black-Scholes-Merton - Preço
    def black_scholes_call_put(S, K, T, r, sigma, option_type):
        d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
        d2 = d1 - sigma * math.sqrt(T)
        
        if option_type == 'call':
            option_price = S * norm_cdf(d1) - K * math.exp(-r * T) * norm_cdf(d2)
        elif option_type == 'put':
            option_price = K * math.exp(-r * T) * norm_cdf(-d2) - S * norm_cdf(-d1)
        else:
            raise ValueError("Invalid option type. Use 'call' or 'put'.")
        
        return option_price

    # Função de distribuição cumulativa normal padrão (CDF)
    def norm_cdf(x):
        return (1.0 + math.erf(x / math.sqrt(2.0))) / 2.0

    # Título do aplicativo
    st.subheader('Calculadora Black-Scholes-Merton - Preço da opção')
    st.markdown("""
    Black, Scholes e Merton revolucionaram a análise de opções, fornecendo um arcabouço robusto 
                para entender e precificar riscos financeiros. Sua influência perdura, moldando 
                a maneira como investidores e instituições lidam com a complexidade dos mercados. 
    """)
    st.markdown('---')
    # Organizando os campos de entrada em duas colunas
    col1, col2 = st.columns(2)

    # Entrada dos valores dos parâmetros
    with col1:
        S = st.number_input('Preço do ativo subjacente (S)', min_value=0.0)
        K = st.number_input('Preço de exercício da opção (K)', min_value=0.0)
        T = st.number_input('Vencimento da opção (T), como fração do ano', min_value=0.0)
    with col2:
        r = st.number_input('Taxa de juros anual (r), como fração', min_value=0.0)
        sigma = st.number_input('Volatilidade (Sigma), como fração', min_value=0.0)
        option_type = st.selectbox('Tipo de opção', ['call', 'put'])

    # Botão para calcular o preço da opção
    if st.button('Calcular preço da opção'):
        option_price = black_scholes_call_put(S, K, T, r, sigma, option_type)
        st.write(f'O preço da {option_type} é: ${round(option_price,2)}')

    st.markdown('---')

    # Função para calcular o preço da opção Black-Scholes-Merton
    def black_scholes_call_put(S, K, T, r, sigma, option_type):
        d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
        d2 = d1 - sigma * math.sqrt(T)
        
        if option_type == 'call':
            option_price = S * stats.norm.cdf(d1) - K * math.exp(-r * T) * stats.norm.cdf(d2)
        elif option_type == 'put':
            option_price = K * math.exp(-r * T) * stats.norm.cdf(-d2) - S * stats.norm.cdf(-d1)
        else:
            raise ValueError("Invalid option type. Use 'call' or 'put'.")
        
        return option_price

    # Função para calcular a volatilidade implícita
    def implied_volatility(option_price, S, K, T, r, option_type):
        # Função para encontrar a volatilidade implícita usando o método de Newton-Raphson
        def func(sigma, option_price, S, K, T, r, option_type):
            return black_scholes_call_put(S, K, T, r, sigma, option_type) - option_price

        # Chamada para o método de Newton-Raphson para encontrar a volatilidade implícita
        sigma = newton(func, x0=0.2, args=(option_price, S, K, T, r, option_type))
        return sigma

    # Título do aplicativo
    st.subheader('Calculadora Black-Scholes-Merton - Volatilidade implícita')
    st.markdown("""
    A volatilidade implícita, elemento-chave no modelo Black-Scholes, reflete as expectativas 
                do mercado sobre a flutuação futura dos preços de ativos. Sua análise é crucial 
                para precificar opções e compreender as percepções dos investidores em relação 
                ao risco. 
    """)
    # Organizando os campos de entrada em duas colunas
    col1, col2 = st.columns(2)

    # Entrada dos valores dos parâmetros
    with col1:
        Sv = st.number_input('Preço do ativo subjacente (S)', min_value=0.0, key='Sv')
        Kv = st.number_input('Preço de exercício da opção (K)', min_value=0.0, key='Kv')
        Tv = st.number_input('Vencimento da opção (T), como fração do ano', min_value=0.0, key='Tv')
    with col2:
        rv = st.number_input('Taxa de juros anual (r), como fração', min_value=0.0, key='rv')
        option_price = st.number_input('Preço da opção', min_value=0.0, key='option_price')
        option_type = st.selectbox('Tipo de opção', ['call', 'put'], key='option_type')

    # Botão para calcular a volatilidade implícita
    if st.button('Calcular volatilidade implícita'):
        sigma_impl = implied_volatility(option_price, Sv, Kv, Tv, rv, option_type)
        st.write(f'A volatilidade implícita é {round(sigma_impl,2)}')

###########################
### GREGAS

elif selected_calculator == "Calculadora de Gregas de Opções":
    # Calculadora de gregas
    # Função para calcular a grega Delta
    def delta(S, K, T, r, sigma, option_type):
        d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))

        if option_type == 'call':
            delta = norm.cdf(d1)
        elif option_type == 'put':
            delta = norm.cdf(d1) - 1
        else:
            raise ValueError("Invalid option type. Use 'call' or 'put'.")
        
        return delta

    # Função para calcular a grega Gamma
    def gamma(S, K, T, r, sigma):
        d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
        gamma = norm.pdf(d1) / (S * sigma * math.sqrt(T))
        return gamma

    # Função para calcular a grega Vega
    def vega(S, K, T, r, sigma):
        d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
        vega = S * norm.pdf(d1) * math.sqrt(T)
        return vega

    # Função para calcular a grega Theta
    def theta(S, K, T, r, sigma, option_type):
        d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
        d2 = d1 - sigma * math.sqrt(T)

        if option_type == 'call':
            theta = -(S * norm.pdf(d1) * sigma) / (2 * math.sqrt(T)) - r * K * math.exp(-r * T) * norm.cdf(d2)
        elif option_type == 'put':
            theta = -(S * norm.pdf(d1) * sigma) / (2 * math.sqrt(T)) + r * K * math.exp(-r * T) * norm.cdf(-d2)
        else:
            raise ValueError("Invalid option type. Use 'call' or 'put'.")
        
        return theta

    # Função para calcular a grega Rho
    def rho(S, K, T, r, sigma, option_type):
        d2 = (math.log(S / K) + (r - 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))

        if option_type == 'call':
            rho = K * T * math.exp(-r * T) * norm.cdf(d2)
        elif option_type == 'put':
            rho = -K * T * math.exp(-r * T) * norm.cdf(-d2)
        else:
            raise ValueError("Invalid option type. Use 'call' or 'put'.")
        
        return rho

    # Título do aplicativo
    st.subheader('Calculadora de Gregas de Opções')
    st.markdown("""
        As 'Gregas', na análise de derivativos, referem-se a medidas de sensibilidade de preço 
                de opções a diferentes variáveis, como volatilidade, tempo e  movimento do preço 
                do ativo subjacente. Delta, Gamma, Vega, Theta e Rho são alguns exemplos essenciais. 
        """)
    st.markdown('---')
    # Organizando os campos de entrada em duas colunas
    col1, col2 = st.columns(2)

    # Entrada dos valores dos parâmetros
    with col1:
        S = st.number_input('Preço do ativo subjacente (S)', min_value=0.0, step=0.01, key='S_gregas')
        K = st.number_input('Preço de exercício da opção (K)', min_value=0.0, step=0.01, key='K_gregas')
        T = st.number_input('Tempo até o vencimento (T) em anos', min_value=0.0, step=0.01, key='T_gregas')
    with col2:
        r = st.number_input('Taxa de juros anual (r), como fração', min_value=0.0, step=0.0001, key='r_gregas')
        sigma = st.number_input('Volatilidade (Sigma), como fração', min_value=0.0, step=0.0001, key='sigma_gregas')
        option_type = st.selectbox('Tipo de opção', ['call', 'put'], key='option_type_gregas')

    # Botão para calcular as "gregas"
    if st.button('Calcular Gregas'):
        st.write("Delta:", round(delta(S, K, T, r, sigma, option_type),6))
        st.write("Gamma:", round(gamma(S, K, T, r, sigma),6))
        st.write("Vega:", round(vega(S, K, T, r, sigma),6))
        st.write("Theta:", round(theta(S, K, T, r, sigma, option_type),6))
        st.write("Rho:", round(rho(S, K, T, r, sigma, option_type),6))

###########################
### CONES DE VOLATILIDADE

elif selected_calculator == "Cones de Volatilidade":
    # Cones de volatilidade para diferentes ativos
    st.subheader('Cones de Volatilidade')
    st.markdown("""
        A parte mais difícil da negociação de opções é determinar se elas estão baratas ou caras. 
                Ao comprar ou vender uma opção, você está exposto à volatilidade do ativo subjacente. 
                Por isso, é importante comparar a volatilidade aos seus níveis recentes. 
                Os cones de volatilidade podem ajudar nessa análise. Veja abaixo gráficos do cone de
                volatilidade para diferentes ativos subjacentes.
        """)
    st.markdown('---')
    acaocone = st.radio('Escolha o ativo subjacente', ['ABEV3','BBDC4','BOVA11','PETR4','VALE3'])    
    windows = [15, 30, 45, 60, 75, 90, 105, 120]
    quantiles = [0.25, 0.75]
    min_ = []
    max_ = []
    median = []
    top_q = []
    bottom_q = []
    realized = []
    start = "2006-01-02"
    def realized_vol(price_data, window=30):
        log_return = (price_data["Close"] / price_data["Close"].shift(1)).apply(np.log)
        return log_return.rolling(window=window, center=False).std() * math.sqrt(252)

    if acaocone == 'ABEV3':
        data = yf.download('ABEV3.SA', start=start)
    if acaocone == 'BBDC4':
        data = yf.download('BBDC4.SA', start=start)
    if acaocone == 'BOVA11':
        data = yf.download('BOVA11.SA', start=start)
    if acaocone == 'PETR4':
        data = yf.download('PETR4.SA', start=start)
    if acaocone == 'VALE3':
        data = yf.download('VALE3.SA', start=start)

    for window in windows:
        # get a dataframe with realized volatility
        estimator = realized_vol(window=window, price_data=data)
        # append the summary stats to a list
        min_.append(estimator.min())
        max_.append(estimator.max())
        median.append(estimator.median())
        top_q.append(estimator.quantile(quantiles[1]))
        bottom_q.append(estimator.quantile(quantiles[0]))
        realized.append(estimator.iloc[-1])
    
    data = [
    go.Scatter(x=windows, y=min_, mode='markers+lines', name='Min'),
    go.Scatter(x=windows, y=max_, mode='markers+lines', name='Max'),
    go.Scatter(x=windows, y=median, mode='markers+lines', name='Mediana'),
    go.Scatter(x=windows, y=top_q, mode='markers+lines', name=f'{quantiles[1] * 100:.0f} Percentil'),
    go.Scatter(x=windows, y=bottom_q, mode='markers+lines', name=f'{quantiles[0] * 100:.0f} Percentil'),
    go.Scatter(x=windows, y=realized, mode='markers+lines', name='Realizado', marker=dict(color='yellow'))
    ]

    # Criar o layout do gráfico
    layout = go.Layout(
        title=f'Cone de Volatilidade - {acaocone}',
        xaxis=dict(title='Janelas'),
        yaxis=dict(title='Valores'),
        legend=dict(x=0.5, y=1.0, bgcolor='rgba(255, 255, 255, 0)', bordercolor='rgba(255, 255, 255, 0)')
    )

    # Criar o gráfico
    fig = go.Figure(data=data, layout=layout)
    st.plotly_chart(fig)

###########################
### PCR - PUT CALL RATIO
    
elif selected_calculator == "PCR - Put Call Ratio":
    # Título do aplicativo
    st.subheader('Put Call Ratio - PCR')
    st.markdown("""
        O Put Call Ratio - PCR é um indicador utilizado para avaliar o sentimento 
                do mercado em relação às opções. Ele compara o volume de negociação 
                de opções de venda com o volume de negociação de opções de compra, 
                oferecendo insights sobre as expectativas dos investidores.
                Alta do PCR pode sinalizar pessimismo em relação ao preço do ativo subjacente.
                Queda do PCR pode sinalizar otimismo em relação ao preço do ativo subjacente.
                Variação indicada abaixo se dá em relação ao valor observado no penúltimo dia útil. 
        """)
    st.markdown('---')
    
    # Carregar as planilhas
    workbook_abev = load_workbook(filename='pcrabev.xlsx')
    sheet_abev = workbook_abev.active  # Ou você pode selecionar uma planilha específica pelo nome: workbook['nome_da_planilha']
    total_rows_abev = sheet_abev.max_row
    ultn_abev = sheet_abev.cell(row=total_rows_abev, column=2).value
    ultv_abev = sheet_abev.cell(row=total_rows_abev, column=3).value
    pultn_abev = sheet_abev.cell(row=(total_rows_abev-1), column=2).value
    pultv_abev = sheet_abev.cell(row=(total_rows_abev-1), column=3).value
    deltan_abev = (ultn_abev/pultn_abev - 1) * 100
    deltav_abev = (ultv_abev/pultv_abev - 1) * 100

    workbook_bbdc = load_workbook(filename='pcrbbdc.xlsx')
    sheet_bbdc = workbook_bbdc.active  # Ou você pode selecionar uma planilha específica pelo nome: workbook['nome_da_planilha']
    total_rows_bbdc = sheet_bbdc.max_row
    ultn_bbdc = sheet_bbdc.cell(row=total_rows_bbdc, column=2).value
    ultv_bbdc = sheet_bbdc.cell(row=total_rows_bbdc, column=3).value
    pultn_bbdc = sheet_bbdc.cell(row=(total_rows_bbdc-1), column=2).value
    pultv_bbdc = sheet_bbdc.cell(row=(total_rows_bbdc-1), column=3).value
    deltan_bbdc = (ultn_bbdc/pultn_bbdc - 1) * 100
    deltav_bbdc = (ultv_bbdc/pultv_bbdc - 1) * 100

    workbook_bova = load_workbook(filename='pcrbova.xlsx')
    sheet_bova = workbook_bova.active  # Ou você pode selecionar uma planilha específica pelo nome: workbook['nome_da_planilha']
    total_rows_bova = sheet_bova.max_row
    ultn_bova = sheet_bova.cell(row=total_rows_bova, column=2).value
    ultv_bova = sheet_bova.cell(row=total_rows_bova, column=3).value
    pultn_bova = sheet_bova.cell(row=(total_rows_bova-1), column=2).value
    pultv_bova = sheet_bova.cell(row=(total_rows_bova-1), column=3).value
    deltan_bova = (ultn_bova/pultn_bova - 1) * 100
    deltav_bova = (ultv_bova/pultv_bova - 1) * 100

    workbook_petr = load_workbook(filename='pcrpetr.xlsx')
    sheet_petr = workbook_petr.active  # Ou você pode selecionar uma planilha específica pelo nome: workbook['nome_da_planilha']
    total_rows_petr = sheet_petr.max_row
    ultn_petr = sheet_petr.cell(row=total_rows_petr, column=2).value
    ultv_petr = sheet_petr.cell(row=total_rows_petr, column=3).value
    pultn_petr = sheet_petr.cell(row=(total_rows_petr-1), column=2).value
    pultv_petr = sheet_petr.cell(row=(total_rows_petr-1), column=3).value
    deltan_petr = (ultn_petr/pultn_petr - 1) * 100
    deltav_petr = (ultv_petr/pultv_petr - 1) * 100

    workbook_vale = load_workbook(filename='pcrvale.xlsx')
    sheet_vale = workbook_vale.active  # Ou você pode selecionar uma planilha específica pelo nome: workbook['nome_da_planilha']
    total_rows_vale = sheet_vale.max_row
    ultn_vale = sheet_vale.cell(row=total_rows_vale, column=2).value
    ultv_vale = sheet_vale.cell(row=total_rows_vale, column=3).value
    pultn_vale = sheet_vale.cell(row=(total_rows_vale-1), column=2).value
    pultv_vale = sheet_vale.cell(row=(total_rows_vale-1), column=3).value
    deltan_vale = (ultn_vale/pultn_vale - 1) * 100
    deltav_vale = (ultv_vale/pultv_vale - 1) * 100

    st.write('##### PCR - Cálculo por número de negócios')
    col1, col2, col3, col4, col5 = st.columns(5)
    # Adicione conteúdo em cada coluna
    with col1:
        st.metric('PCR ABEV3', value=round(ultn_abev,2), delta=f'{round(deltan_abev,2)}%')

    with col2:
        st.metric('PCR BBDC4', value=round(ultn_bbdc,2), delta=f'{round(deltan_bbdc,2)}%')

    with col3:
        st.metric('PCR BOVA11', value=round(ultn_bova,2), delta=f'{round(deltan_bova,2)}%')

    with col4:
        st.metric('PCR PETR4', value=round(ultn_petr,2), delta=f'{round(deltan_petr,2)}%')

    with col5:
        st.metric('PCR VALE3', value=round(ultn_vale,2), delta=f'{round(deltan_vale,2)}%')
    
    st.markdown('---')

    st.write('##### PCR - Cálculo por volume negociado')
    col1, col2, col3, col4, col5 = st.columns(5)
    # Adicione conteúdo em cada coluna
    with col1:
        st.metric('PCR ABEV3', value=round(ultv_abev,2), delta=f'{round(deltav_abev,2)}%')

    with col2:
        st.metric('PCR BBDC4', value=round(ultv_bbdc,2), delta=f'{round(deltav_bbdc,2)}%')

    with col3:
        st.metric('PCR BOVA11', value=round(ultv_bova,2), delta=f'{round(deltav_bova,2)}%')

    with col4:
        st.metric('PCR PETR4', value=round(ultv_petr,2), delta=f'{round(deltav_petr,2)}%')

    with col5:
        st.metric('PCR VALE3', value=round(ultv_vale,2), delta=f'{round(deltav_vale,2)}%')

###########################
### Seguro da Carteira
    
elif selected_calculator == "Seguro da Carteira":
    # Título do aplicativo
    st.subheader('Seguro da Carteira')
    st.markdown("""
        Proteger sua carteira contra disaster risk é essencial. A estratégia de compra de puts 
                fora do dinheiro oferece uma camada de segurança adicional, minimizando exposições 
                indesejadas e fortalecendo a resiliência de seus investimentos. Uma maneira de 
                operacionalizar essa estratégia de proteção consiste em alocar um pequeno percentual 
                da sua carteira em puts de BOVA11 bem fora do dinheiro. Abaixo, algumas candidatas
                (tabela atualizada diariamente). 
        """)
    st.markdown('---')

    ult_bova11_disaster = yf.download('BOVA11.SA')['Close'].iloc[-1]
    # Carregar a planilha
    bova11_disaster = pd.read_excel('bova11_disaster.xlsx')
    bova11_disaster['Fração do Spot'] = bova11_disaster['Strike'] / ult_bova11_disaster
    bova11_disaster = bova11_disaster.drop(columns=['Subjacente', 'Tipo'])
    bova11_disaster['Volume'] = bova11_disaster['Volume'].round(2)
    bova11_disaster = bova11_disaster.round(2)
    html = bova11_disaster.to_html(index=False)
    st.write(html, unsafe_allow_html=True)
    
###########################
### Top Fundos Quantitativos
    
elif selected_calculator == "Top 10 Fundos Quantitativos":
    # Título do aplicativo
    st.subheader('Top 10 Fundos "Quantitativos"')
    st.markdown("""
        Fundos quantitativos revolucionaram a gestão de investimentos ao utilizar algoritmos 
                avançados e análise de dados para tomar decisões. Combinando matemática, 
                estatística e tecnologia, esses fundos buscam maximizar retornos e mitigar 
                riscos de forma inovadora no mercado financeiro. Abaixo, listamos os dez fundos 
                no Brasil com o termo "quant" na denominação com maior retorno dentro do 
                mês corrente. 
        """)
    st.markdown('---')

    # Fetch - MUDAR OS DOIS quando virar o mês
    arquivo = 'inf_diario_fi_202404.csv'
    link = 'https://dados.cvm.gov.br/dados/FI/DOC/INF_DIARIO/DADOS/inf_diario_fi_202404.zip'

    r = requests.get(link)
    zf = zipfile.ZipFile(io.BytesIO(r.content))

    arquivo_fi = zf.open(arquivo)
    linhas = arquivo_fi.readlines()
    linhas = [i.strip().decode('ISO-8859-1') for i in linhas]
    linhas = [i.split(';') for i in linhas]
    df = pd.DataFrame(linhas, columns=linhas[0])
    informes_diarios = df[1:].reset_index()
    informes_diarios[['VL_TOTAL', 'VL_QUOTA', 'VL_PATRIM_LIQ', 'CAPTC_DIA', 'RESG_DIA', 'NR_COTST']] = informes_diarios[
        ['VL_TOTAL', 'VL_QUOTA', 'VL_PATRIM_LIQ', 'CAPTC_DIA', 'RESG_DIA', 'NR_COTST']].apply(pd.to_numeric)
    # Dados Cadastrais
    url = 'http://dados.cvm.gov.br/dados/FI/CAD/DADOS/cad_fi.csv'
    cadastral = pd.read_csv(url, sep=';', encoding='ISO-8859-1')
    # Filtrar fundos com "quant" na denominação
    selecao = cadastral[cadastral['DENOM_SOCIAL'].str.contains('quant', case=False)]
    selecao = selecao[selecao['SIT'] != 'CANCELADA']
    # Filtrar dados de fundos selecionados
    fundos_selecionados = informes_diarios[informes_diarios['CNPJ_FUNDO'].isin(selecao['CNPJ_FUNDO'])]
    # Calcular maiores altas apenas entre fundos selecionados
    retornos = fundos_selecionados.pivot(index='DT_COMPTC', columns='CNPJ_FUNDO', values='VL_QUOTA')
    retornos = (retornos / retornos.iloc[1] - 1) * 100
    maiores_altas = retornos.iloc[-2].sort_values(ascending=False)
    # Selecionar as 10 maiores altas
    top_10_altas = maiores_altas.head(10)
    # Exibir as 10 maiores altas
    print("10 Maiores Altas para Fundos com 'quant' na Denominação:")
    print(top_10_altas)
    # Criar DataFrame com as 10 maiores altas e a denominação social
    top_10_df = pd.DataFrame({'CNPJ_FUNDO': top_10_altas.index, 'Retorno (%)': top_10_altas.values})
    top_10_df['Denominação Social'] = [cadastral[cadastral['CNPJ_FUNDO'] == cnpj]['DENOM_SOCIAL'].values[0] for cnpj in top_10_df['CNPJ_FUNDO']]
    # Indices de 1 a 10
    top_10_df.index = range(1, 11)
    # Formatar a coluna "Retorno (%)" com duas casas decimais
    top_10_df['Retorno (%)'] = top_10_df['Retorno (%)'].round(2)
    top_10_df = top_10_df[['Denominação Social', 'Retorno (%)']]

    st.dataframe(top_10_df)

################################
### Carteiras
    
elif selected_calculator == "Carteiras":
    # Título do aplicativo
    st.subheader('Carteiras')
    st.markdown("""
        O Factor Investing é uma estratégia que busca capturar retornos excedentes ao mirar 
                em fatores específicos, como valor, momento, tamanho, qualidade, baixa 
                volatilidade, aderência a padrões esperados em ESG e outras características 
                dentro de uma carteira diversificada. Compreender esses fatores e suas 
                interações é crucial para decisões de investimento. 
        """)
    st.markdown("""
        A carteira por risk parity é uma estratégia de alocação de ativos que busca 
                equalizar o risco de cada componente da carteira. Isso significa que, 
                em uma carteira por risk parity, cada ativo contribui de forma igual 
                para a volatilidade total da carteira. Essa equalização do risco pode 
                ser alcançada através da atribuição de pesos diferentes para cada ativo, 
                levando em consideração suas correlações e volatilidades históricas. 
        """)
    st.markdown('---')

    # Abrir colunas para selecionar as carteiras em radio
    acaocarteiras = st.radio('Escolha a carteira por critério', ['Magic Formula de Joel Greenblatt', "Risk Parity"])

    if acaocarteiras == 'Magic Formula de Joel Greenblatt':
        # Codigo MF
        dfraw = fundamentus.get_resultado_raw()
        df = fundamentus.get_resultado()
        # Primeiros filtros
        df2 = df[df.pl > 0]
        df3 = df2[df2.evebit > 0]
        df4 = df3[df3.roic > 0]
        df5 = df4[df4.patrliq > 100000000]
        stocks_df = df5[df5.liq2m > 0]
        stocks_df = stocks_df[['evebit', 'roic']]
        # Trabalhando apenas com dy e roic
        data = {'Stock': stocks_df.index,
                'Earnings_Yield': stocks_df['evebit'],
                'ROIC': stocks_df['roic']}
        stocks_df = pd.DataFrame(data)
        # Ordenando dy e roic
        stocks_df['Earnings_Yield_Rank'] = stocks_df['Earnings_Yield'].rank(ascending=True)
        stocks_df['ROIC_Rank'] = stocks_df['ROIC'].rank(ascending=False)
        # Calculando a magic formula
        stocks_df['Magic_Formula_Rank'] = stocks_df['Earnings_Yield_Rank'] + stocks_df['ROIC_Rank']
        # Ordenando pela magic formula
        sorted_stocks = stocks_df.sort_values('Magic_Formula_Rank')
        # Visualizando
        ativos = sorted_stocks.head(30)
        # Removendo baixa liquidez
        codigos_a_remover = ['CEDO4', 'RSUL4', 'CEDO3', 'CAMB3', 'PETR3', 'MTSA4', 'DEXP3', 'MRSA6B']
        ativos = ativos.drop(codigos_a_remover, axis=0)
        carteiramf = ativos.head(10)
        colunas_para_remover = ['Earnings_Yield_Rank', 'ROIC_Rank', 'Magic_Formula_Rank']
        carteiramf = carteiramf.drop(columns=colunas_para_remover)
        del carteiramf['Stock']
        novos_nomes_colunas = {'Earnings_Yield': 'Earnings Yield', 'ROIC': 'ROIC'}
        carteiramf = carteiramf.rename(columns=novos_nomes_colunas)
        carteiramf = carteiramf.rename_axis('Código', axis='index')
        st.markdown('---')
        st.markdown("""
            A estratégia Magic Formula de investimento, popularizada por Joel Greenblatt, 
                    busca empresas com baixa avaliação e alta rentabilidade. Baseia-se 
                    em dois critérios: retorno sobre o capital e múltiplo preço/lucro. 
                    Investidores usam essa fórmula para encontrar ações com potencial 
                    de longo prazo. Na seleção abaixo, nossa adaptação prevê pesos iguais
                    na carteira, ou seja, 10% em cada ação.  
            """)
        st.dataframe(carteiramf)

    if acaocarteiras == 'Risk Parity':
        st.markdown('---')
        st.markdown("""
        Risk Parity é uma abordagem de alocação de ativos que equilibra os riscos entre 
                    diferentes classes de ativos. Diferente da tradicional alocação 
                    baseada em pesos, o Risk Parity considera a volatilidade e a correlação
                    como drivers preponderantes. Isso pode ajudar a diversificar o 
                    risco e potencialmente melhorar os retornos da carteira. 
        """)
        # Lista das ações
        assets = [
            "ABEV3.SA", "ALPA4.SA", "ARZZ3.SA", "ASAI3.SA", "AZUL4.SA", "B3SA3.SA", "BBAS3.SA", 
            "BBDC3.SA", "BBDC4.SA", "BBSE3.SA", "BEEF3.SA", "BPAC11.SA", "BRAP4.SA", "BRFS3.SA", "BRKM5.SA", 
            "CASH3.SA", "CCRO3.SA", "CIEL3.SA", "CMIG4.SA", "CMIN3.SA", "COGN3.SA", "CPFE3.SA", "CPLE6.SA", 
            "CRFB3.SA", "CSAN3.SA", "CSNA3.SA", "CVCB3.SA", "CYRE3.SA", "DXCO3.SA", "EGIE3.SA", "ELET3.SA", 
            "ELET6.SA", "EMBR3.SA", "ENEV3.SA", "ENGI11.SA", "EQTL3.SA", "EZTC3.SA", "FLRY3.SA", 
            "GGBR4.SA", "GOAU4.SA", "GOLL4.SA", "HAPV3.SA", "HYPE3.SA", "IGTI11.SA", "IRBR3.SA", "ITSA4.SA", 
            "ITUB4.SA", "JBSS3.SA", "KLBN11.SA", "LREN3.SA", "LWSA3.SA", "MGLU3.SA", "MRFG3.SA", "MRVE3.SA", 
            "MULT3.SA", "NTCO3.SA", "PCAR3.SA", "PETR4.SA", "PETZ3.SA", "PRIO3.SA", "RADL3.SA", 
            "RAIL3.SA", "RAIZ4.SA", "RDOR3.SA", "RENT3.SA", "RRRP3.SA", "SANB11.SA", "SBSP3.SA", "SLCE3.SA", 
            "SMTO3.SA", "SOMA3.SA", "SUZB3.SA", "TAEE11.SA", "TIMS3.SA", "TOTS3.SA", "UGPA3.SA", "USIM5.SA", 
            "VALE3.SA", "VBBR3.SA", "BHIA3.SA", "VIVT3.SA", "WEGE3.SA", "YDUQ3.SA"
        ]

        #download data
        end = datetime.now()
        start = end - timedelta(days = 180)
        data = yf.download(assets, start=start, end=end)
        # compute non-compounding, daily returns
        returns = data['Adj Close'].pct_change().dropna()

        # Portfolio with equal risk contribution weights
        port = rp.Portfolio(returns=returns)
        port.assets_stats(method_mu='hist', method_cov='hist', d=0.94)
        w_rp = port.rp_optimization(
            model="Classic",  # use historical
            rm="MV",  # use mean-variance optimization
            hist=True,  # use historical scenarios
            rf=0,  # set risk free rate to 0
            b=None  # don't use constraints
        )

        # Portfolio with minimum return constraint
        port.lowerret = 0.0019
        # estimate the optimal portfolio with risk parity with the constraint
        w_rp_c = port.rp_optimization(
            model="Classic",  # use historical
            rm="MV",  # use mean-variance optimization
            hist=True,  # use historical scenarios
            rf=0,  # set risk free rate to 0
            b=None  # don't use constraints
        )

        dfrp_ordenado = w_rp_c.sort_values(by='weights', ascending=False)
        dfrp = dfrp_ordenado.head(10)
        dfrp.index = dfrp.index.astype(str).str.replace('.SA', '')
        total = dfrp['weights'].sum()
        dfrp['Pesos (%)'] = round((dfrp['weights'] / total) * 100,2)
        dfrp = dfrp.drop(columns=['weights'])
        dfrp = dfrp.rename_axis('Código', axis='index')
        st.dataframe(dfrp)


################################
### BDR - Spreads
    
elif selected_calculator == "BDR - Spreads":
    # Título do aplicativo
    st.subheader('BDR - Maiores e Menores Spreads')
    st.markdown("""
        O spread entre ações americanas e BDRs reflete as diferenças de preços e 
                condições de mercado entre os EUA e o Brasil. Enquanto as ações americanas 
                são negociadas nos mercados dos EUA, os BDRs são recibos de ações estrangeiras 
                negociadas na Bolsa brasileira. Há um spread, uma diferença entre o valor da ação
                 e o da BDR equivalente (lembrando que um BDR não necessariamente corresponde a 
                apenas um papel). Para calcular, você deve multiplicar o preço da ação pelo câmbio 
                atual e pela proporção de ações no lote. A diferença entre esse valor e o do BDR é 
                o spread. O spread pode variar devido a fatores como taxas de câmbio, liquidez 
                e regulamentações. Os investidores devem estar atentos a essas diferenças ao 
                considerar suas estratégias de investimento em ações globais. Abaixo, os maiores
                e menores valores de spread para BDRs integrantes do índice BDRX B3.
        """)
    st.markdown('---')

    bdrmaiores = pd.read_excel('bdrmaiores.xlsx')
    bdrmenores = pd.read_excel('bdrmenores.xlsx')

    # Exibir o DataFrame
    st.markdown('#### Maiores Spreads')
    st.markdown(bdrmaiores.style.hide(axis="index").to_html(), unsafe_allow_html=True)
    st.markdown('---')
    st.markdown('#### Menores Spreads')
    st.markdown(bdrmenores.style.hide(axis="index").to_html(), unsafe_allow_html=True)

################################
### Long Short - Cointegração
    
elif selected_calculator == "Long Short - Cointegração":
    # Título do aplicativo
    st.subheader('Long Short - Cointegração')
    st.markdown("""
        A estratégia de long short com cointegração e o cálculo de reversão à média de 
                Ornstein-Uhlenbeck são ferramentas elegíveis na seleção de pares 
                no mercado de capitais. A cointegração é uma relação estatística entre 
                duas séries temporais que indica que, mesmo que elas se movam independentemente 
                no curto prazo, elas têm uma relação de longo prazo estável. 
                Essa estabilidade permite identificar pares de ativos que têm uma relação de 
                equilíbrio e, portanto, podem ser utilizados em estratégias de long short. Nessa 
                estratégia de long short, um par de ativos é selecionado com base na cointegração. 
                Um dos ativos é comprado (posição longa) enquanto o outro é vendido (posição short). 
                Isso é feito na expectativa de que, embora os preços dos ativos possam se 
                afastar temporariamente de seu equilíbrio, eles eventualmente retornarão 
                a esse equilíbrio de longo prazo, gerando lucros para o investidor.
                O cálculo de reversão à média de Ornstein-Uhlenbeck é uma técnica que pode ser usada 
                para estimar o tempo que leva para que os preços de um par de ativos voltem ao 
                equilíbrio após um desvio. Isso é importante para determinar o momento ideal 
                para entrar e sair de uma posição, maximizando os lucros potenciais da estratégia 
                de long short.

        """)
    st.markdown('---')
    end = datetime.now()
    start = end - timedelta(days = 200)
    #from ibov composition file
    assets = ["ABEV3.SA", "ALPA4.SA", "ARZZ3.SA", "ASAI3.SA", "AZUL4.SA", "B3SA3.SA", "BBAS3.SA", 
            "BBDC3.SA", "BBDC4.SA", "BBSE3.SA", "BEEF3.SA", "BPAC11.SA", "BRAP4.SA", "BRFS3.SA", "BRKM5.SA", 
            "CASH3.SA", "CCRO3.SA", "CIEL3.SA", "CMIG4.SA", "CMIN3.SA", "COGN3.SA", "CPFE3.SA", "CPLE6.SA", 
            "CRFB3.SA", "CSAN3.SA", "CSNA3.SA", "CVCB3.SA", "CYRE3.SA", "DXCO3.SA", "EGIE3.SA", "ELET3.SA", 
            "ELET6.SA", "EMBR3.SA", "ENEV3.SA", "ENGI11.SA", "EQTL3.SA", "EZTC3.SA", "FLRY3.SA", 
            "GGBR4.SA", "GOAU4.SA", "GOLL4.SA", "HAPV3.SA", "HYPE3.SA", "IGTI11.SA", "IRBR3.SA", "ITSA4.SA", 
            "ITUB4.SA", "JBSS3.SA", "KLBN11.SA", "LREN3.SA", "LWSA3.SA", "MGLU3.SA", "MRFG3.SA", "MRVE3.SA", 
            "MULT3.SA", "NTCO3.SA", "PCAR3.SA", "PETR3.SA", "PETR4.SA", "PETZ3.SA", "PRIO3.SA", "RADL3.SA", 
            "RAIL3.SA", "RAIZ4.SA", "RDOR3.SA", "RENT3.SA", "RRRP3.SA", "SANB11.SA", "SBSP3.SA", "SLCE3.SA", 
            "SMTO3.SA", "SOMA3.SA", "SUZB3.SA", "TAEE11.SA", "TIMS3.SA", "TOTS3.SA", "UGPA3.SA", "USIM5.SA", 
            "VALE3.SA", "VBBR3.SA", "BHIA3.SA", "VIVT3.SA", "WEGE3.SA", "YDUQ3.SA"]
    quotes = yf.download(assets, start = start, end = end)["Adj Close"]
    #drop a column
    quotes = quotes.drop(quotes.columns[1], axis=1)
    quotes.isna().sum().sum()  # Checking for NAs
    # Remove '.SA'
    quotes.columns = [col[:-3] for col in quotes.columns]

    # Exibir o DataFrame
    st.markdown('#### 10 Pares considerando dados em D-1')
    st.markdown('---')
    # Carregar os pares do arquivo XLSX
    pairs_df = pd.read_excel("tenpairs.xlsx")
  
    # Iterando sobre as 10 primeiras linhas do DataFrame
    for index, row in pairs_df.head(10).iterrows():
    # Extrair o par de ações
        acao1 = row['Acao1']
        acao2 = row['Acao2']

        # Extraindo os dados para o par
        data1 = quotes[acao1]
        data2 = quotes[acao2]

        # Criando o gráfico
        fig, ax1 = plt.subplots(figsize=(10, 6))

        # Plotando a primeira série no lado esquerdo
        color = 'tab:red'
        ax1.set_xlabel('Data')
        ax1.set_ylabel(acao1, color=color)
        ax1.plot(data1.index, data1.values, color=color)
        ax1.tick_params(axis='y', labelcolor=color)

        # Criando o segundo eixo y para a segunda série no lado direito
        ax2 = ax1.twinx()
        color = 'tab:blue'
        ax2.set_ylabel(acao2, color=color)
        ax2.plot(data2.index, data2.values, color=color)
        ax2.tick_params(axis='y', labelcolor=color)

        # Configurando títulos dos eixos e legendas
        plt.title(f'{acao1} vs {acao2}')

        # Exibindo o gráfico no Streamlit
        st.pyplot(fig)



################################
### Monitor de 5 Dias
    
elif selected_calculator == "Monitor de 5 Dias":
    # Título do aplicativo
    st.subheader('Monitor de 5 Dias')
    st.markdown("""
        Visão do mercado de ações nos últimos 5 dias úteis. Esta ferramenta destaca as 10 ações 
                que mais subiram e mais caíram em valor, permitindo identificação das tendências 
                de mercado. Além disso, fornece informações sobre as 10 ações com maior alta 
                e maior baixa de volume médio de negociação, oferecendo insights sobre o 
                interesse dos investidores em determinados ativos. Por fim, destaca as 10 ações 
                com maior alta e maior baixa de volatilidade média, destacando os movimentos 
                de preço mais significativos e potencialmente indicando oportunidades de negociação.                 
        """)
    st.markdown('---')

    # Datas
    end = datetime.now()
    start = end - timedelta(days=60)

    # fetch
    ativos = ["ABEV3.SA", "ALPA4.SA", "ARZZ3.SA", "ASAI3.SA", "AZUL4.SA", "B3SA3.SA", "BBAS3.SA", 
            "BBDC3.SA", "BBDC4.SA", "BBSE3.SA", "BEEF3.SA", "BPAC11.SA", "BRAP4.SA", "BRFS3.SA", "BRKM5.SA", 
            "CASH3.SA", "CCRO3.SA", "CIEL3.SA", "CMIG4.SA", "CMIN3.SA", "COGN3.SA", "CPFE3.SA", "CPLE6.SA", 
            "CRFB3.SA", "CSAN3.SA", "CSNA3.SA", "CVCB3.SA", "CYRE3.SA", "DXCO3.SA", "EGIE3.SA", "ELET3.SA", 
            "ELET6.SA", "EMBR3.SA", "ENEV3.SA", "ENGI11.SA", "EQTL3.SA", "EZTC3.SA", "FLRY3.SA", 
            "GGBR4.SA", "GOAU4.SA", "GOLL4.SA", "HAPV3.SA", "HYPE3.SA", "IGTI11.SA", "IRBR3.SA", "ITSA4.SA", 
            "ITUB4.SA", "JBSS3.SA", "KLBN11.SA", "LREN3.SA", "LWSA3.SA", "MGLU3.SA", "MRFG3.SA", "MRVE3.SA", 
            "MULT3.SA", "NTCO3.SA", "PCAR3.SA", "PETR3.SA", "PETR4.SA", "PETZ3.SA", "PRIO3.SA", "RADL3.SA", 
            "RAIL3.SA", "RAIZ4.SA", "RDOR3.SA", "RENT3.SA", "RRRP3.SA", "SANB11.SA", "SBSP3.SA", "SLCE3.SA", 
            "SMTO3.SA", "SOMA3.SA", "SUZB3.SA", "TAEE11.SA", "TIMS3.SA", "TOTS3.SA", "UGPA3.SA", "USIM5.SA", 
            "VALE3.SA", "VBBR3.SA", "BHIA3.SA", "VIVT3.SA", "WEGE3.SA", "YDUQ3.SA"]

    acaocinco = st.radio('Escolha a tabela', ['Retornos nos últimos 5 dias úteis','Volumes nos últimos 5 dias úteis','Volatilidades nos últimos 5 dias úteis'])

    if acaocinco == 'Retornos nos últimos 5 dias úteis':
        st.markdown('---')
        st.markdown("""
        #### 10 ações que mais subiram e mais caíram nos últimos 5 dias úteis
        """)
        ## Retornos
        cotas = yf.download(ativos, start=start, end=end)["Adj Close"]

        # Calcular os retornos percentuais diários
        retornos_diarios = cotas.pct_change()

        # Calcular a variação percentual nos últimos 5 dias úteis
        var_percentual_5d = retornos_diarios.tail(5).mean()

        # Identificar os 10 maiores e menores retornos
        maiores_retornos = var_percentual_5d.nlargest(10)
        menores_retornos = var_percentual_5d.nsmallest(10)

        df_maiores_retornos = pd.DataFrame({'Ticker': maiores_retornos.index.str.replace('.SA', ''), 
                                            'Retorno (%)': (maiores_retornos.values * 100)})
        df_menores_retornos = pd.DataFrame({'Ticker': menores_retornos.index.str.replace('.SA', ''), 
                                            'Retorno (%)': (menores_retornos.values * 100)})

        df_maiores_retornos['Retorno (%)'] = df_maiores_retornos['Retorno (%)'].apply(lambda x: '{:.2f}'.format(x))
        df_menores_retornos['Retorno (%)'] = df_menores_retornos['Retorno (%)'].apply(lambda x: '{:.2f}'.format(x))
        st.markdown('---')
        st.markdown("""
        ##### 10 maiores altas de preço
        """)
        st.markdown(df_maiores_retornos.style.hide(axis="index").to_html(), unsafe_allow_html=True)
        st.markdown('---')
        st.markdown("""
        ##### 10 maiores quedas de preço
        """)    
        st.markdown(df_menores_retornos.style.hide(axis="index").to_html(), unsafe_allow_html=True)


    if acaocinco == 'Volumes nos últimos 5 dias úteis':
        st.markdown('---')
        st.markdown("""
        #### 10 ações que mais tiveram alta e queda de volume negociado nos últimos 5 dias úteis
        """)
        ## Volume
        volume = yf.download(ativos, start=start, end=end)["Volume"]
        # Calcular os retornos percentuais diários
        mudanca_diaria_volume = volume.pct_change()
        # Calcular a variação percentual média nos últimos 5 dias úteis
        var_percentual_5d_vol = mudanca_diaria_volume.tail(5).mean()

        # Identificar os 10 maiores e menores retornos
        volume_cresce = var_percentual_5d_vol.nlargest(10)
        volume_cai = var_percentual_5d_vol.nsmallest(10)

        df_maiores_altasvol = pd.DataFrame({'Ticker': volume_cresce.index.str.replace('.SA', ''), 
                                            'Mudança (%)': (volume_cresce.values * 100)})
        df_maiores_quedasvol = pd.DataFrame({'Ticker': volume_cai.index.str.replace('.SA', ''), 
                                            'Mudança (%)': (volume_cai.values * 100)})
        
        df_maiores_altasvol['Mudança (%)'] = df_maiores_altasvol['Mudança (%)'].apply(lambda x: '{:.2f}'.format(x))
        df_maiores_quedasvol['Mudança (%)'] = df_maiores_quedasvol['Mudança (%)'].apply(lambda x: '{:.2f}'.format(x))

        st.markdown('---')
        st.markdown("""
        ##### 10 maiores altas de volume
        """)
        st.markdown(df_maiores_altasvol.style.hide(axis="index").to_html(), unsafe_allow_html=True)
        st.markdown('---')
        st.markdown("""
        ##### 10 maiores quedas de volume
        """)    
        st.markdown(df_maiores_quedasvol.style.hide(axis="index").to_html(), unsafe_allow_html=True)
    
    if acaocinco == 'Volatilidades nos últimos 5 dias úteis':
        st.markdown('---')
        st.markdown("""
        #### 10 ações que mais tiveram variação de volatilidade nos últimos 5 dias úteis
        """)
        st.markdown('---')
        # Calcular a volatilidade diária em uma janela móvel dos últimos 22 dias úteis para cada ação
        cotas = yf.download(ativos, start=start, end=end)["Adj Close"]
        volatilidade = cotas.pct_change(fill_method=None).rolling(window=22).std()
        diff_volatilidade = volatilidade.diff(periods=5)
        maiores_crescimentos = diff_volatilidade.max().nlargest(10)
        maiores_quedas = diff_volatilidade.min().nsmallest(10)
        volatilidade.index = volatilidade.index.astype(str).str.replace('.SA', '')
        maiores_crescimentos.index = maiores_crescimentos.index.str.replace('.SA', '')
        maiores_quedas.index = maiores_quedas.index.str.replace('.SA', '')

        nomes_maiores_crescimentos = maiores_crescimentos.index.tolist()
        nomes_maiores_quedas = maiores_quedas.index.tolist()

        # Criar um novo DataFrame combinando os nomes das ações com os títulos das colunas
        df_maiores_crescimentos_quedas = pd.DataFrame({
            "Maiores altas": nomes_maiores_crescimentos,
            "Maiores quedas": nomes_maiores_quedas
        })

        st.markdown(df_maiores_crescimentos_quedas.style.hide(axis="index").to_html(), unsafe_allow_html=True)


elif selected_calculator == "Long Short - Teste seu Par":
    # Título do aplicativo
    st.subheader('Long Short - Teste seu Par por Cointegração')
    st.markdown("""
        A cointegração é uma relação estatística entre duas séries temporais que indica que, 
                mesmo que elas se movam independentemente no curto prazo, elas têm uma 
                relação de longo prazo estável. A identificação de pares cointegrados passa principalmente
                por testes e análises que compreendem manipulações de regressão linear simples, 
                tendência e estacionariedade.  
             """)
    st.markdown('---')
    #from ibov composition file
    end = datetime.now()
    start = end - timedelta(days = 200)

    # Funções para realizar a regressão linear e retornar beta, p-valor do beta e residuos
    def linear_regression(x, y):
        x = sm.add_constant(x)
        model = OLS(y, x).fit()
        return round(model.params[0], 2)  # Retorna o coeficiente beta
    
    def linear_regressionp(x, y):
        x = sm.add_constant(x)
        model = OLS(y, x).fit()
        return model.pvalues.iloc[0]  # Retorna o p-valor do coeficiente beta
    
    def linear_regressionr(x, y):
        x = sm.add_constant(x)
        model = OLS(y, x).fit()
        return model.resid  # Retorna os residuos da regressao
    # Entrada dos códigos das ações
    col1, col2 = st.columns(2)

    # Adicionar conteúdo em cada coluna
    with col1:
        stock_code1 = st.text_input('Digite o código da primeira ação (preferencialmente do Ibovespa):')

    with col2:
        stock_code2 = st.text_input('Digite o código da segunda ação (preferencialmente do Ibovespa):')

    # Botão para enviar os códigos e realizar as análises
    if st.button('Enviar'):
        # Obtendo os dados históricos das ações 
        stock_data1 = yf.download(stock_code1 + '.SA', start = start, end = end)["Adj Close"]
        stock_data2 = yf.download(stock_code2 + '.SA', start = start, end = end)["Adj Close"]
        residuos = linear_regressionr(stock_data2, stock_data1)
        pbeta = linear_regressionp(stock_data2, stock_data1)

        # Verificando se os dados são suficientes para análise
        if len(stock_data1) < 2 or len(stock_data2) < 2:
            st.error('Não há dados suficientes para análise.')
        else:
            # Realizando os testes de raiz unitária
            unit_root_test_result1 = adfuller(stock_data1, autolag='AIC')[1] > 0.1
            unit_root_test_result2 = adfuller(stock_data2, autolag='AIC')[1] > 0.1
            unit_root_test_result3 = adfuller(residuos, autolag='AIC')[1] < 0.1
            beta_test_result = pbeta < 0.1

            # Realizando a regressão linear
            beta = None
            if unit_root_test_result1 and unit_root_test_result2 and unit_root_test_result3 and beta_test_result:
                beta = linear_regression(stock_data1, stock_data2)

            # Exibindo o resultado do beta, se disponível
            if beta is not None:
                st.write(f'O par é cointegrado e o coeficiente de cointegração é: {beta}')
            else:
                st.warning('O par não é cointegrado.')

            # Plotando o gráfico das cotações das ações
            fig, ax1 = plt.subplots()

            color = 'tab:red'
            ax1.set_xlabel('Data')
            ax1.set_ylabel(f'Preço de {stock_code1}', color=color)
            ax1.plot(stock_data1.index, stock_data1, color=color)
            ax1.tick_params(axis='y', labelcolor=color)

            ax2 = ax1.twinx()
            color = 'tab:blue'
            ax2.set_ylabel(f'Preço de {stock_code2}', color=color)
            ax2.plot(stock_data2.index, stock_data2, color=color)
            ax2.tick_params(axis='y', labelcolor=color)

            st.pyplot(fig)