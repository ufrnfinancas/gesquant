# PCR
import pandas as pd
import requests
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

vencimento = '2024-10-18'
# Lista de feriados no Brasil (pode ser necessário atualizar)
feriados_brasil = [
    '2024-01-01',  # Ano Novo
    '2024-02-12',  # Carnaval
    '2024-02-13',  # Carnaval
    '2024-03-29',  # Paixão de Cristo
    '2024-05-01',  # Dia do Trabalho
    '2024-05-30',  # Corpus Christi
    '2024-11-15',  # Proclamação da República
    '2024-11-20',  # Zumbi
    '2024-12-24',  # Natal
    '2024-12-25',   # Natal
    '2024-12-31',  # Ano novo
]

# Converter feriados para formato de data
feriados_brasil = [datetime.strptime(data, '%Y-%m-%d').date() for data in feriados_brasil]

# Verificar se o dia é útil
def eh_dia_util(data):
    return data.weekday() < 5 and data not in feriados_brasil

# Encontrar o último dia útil
data_ontem = datetime.today().date() - timedelta(days=1)
while not eh_dia_util(data_ontem):
    data_ontem -= timedelta(days=1)
data_ontem


### OS CÓDIGOS PARA CADA AÇÃO

subjacente = 'ABEV3'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrabev.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrabev.xlsx')


subjacente = 'BBDC4'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrbbdc.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrbbdc.xlsx')


subjacente = 'BOVA11'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrbova.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrbova.xlsx')


subjacente = 'PETR4'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrpetr.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrpetr.xlsx')


subjacente = 'VALE3'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrvale.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrvale.xlsx')
