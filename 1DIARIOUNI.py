# Pairs Trading B3

import math
import pandas as pd
import numpy as np
from itertools import combinations
from statsmodels.regression.linear_model import OLS
import statsmodels.api as sm
from statsmodels.tsa.stattools import adfuller
from arch.unitroot import PhillipsPerron
import matplotlib.pyplot as plt
import datetime
from datetime import timedelta, time, datetime
import yfinance as yf
import time as tm
from scipy.optimize import minimize
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Datas
end = datetime.now()
start = end - timedelta(days = 200)
#from ibov composition file
assets = ["ABEV3.SA", "ALPA4.SA", "ASAI3.SA", "AZUL4.SA", "B3SA3.SA", "BBAS3.SA", 
        "BBDC3.SA",  "BBDC4.SA", "BBSE3.SA", "BEEF3.SA", "BPAC11.SA", "BRAP4.SA", "BRFS3.SA", "BRKM5.SA", 
        "CASH3.SA", "CCRO3.SA", "CMIG4.SA", "CMIN3.SA", "COGN3.SA", "CPFE3.SA", "CPLE6.SA", 
        "CRFB3.SA", "CSAN3.SA", "CSNA3.SA", "CVCB3.SA", "CYRE3.SA", "DXCO3.SA", "EGIE3.SA", "ELET3.SA", 
        "ELET6.SA", "EMBR3.SA", "ENEV3.SA", "ENGI11.SA", "EZTC3.SA", "FLRY3.SA", 
        "GGBR4.SA", "GOAU4.SA", "GOLL4.SA", "HAPV3.SA", "HYPE3.SA", "IGTI11.SA", "IRBR3.SA", "ITSA4.SA", 
        "ITUB4.SA", "JBSS3.SA", "KLBN11.SA", "LWSA3.SA", "MGLU3.SA", "MRFG3.SA", "MRVE3.SA", 
        "MULT3.SA", "NTCO3.SA", "PCAR3.SA", "PETR3.SA", "PETR4.SA", "PETZ3.SA", "PRIO3.SA", "RADL3.SA", 
        "RAIL3.SA", "RAIZ4.SA", "RDOR3.SA", "RENT3.SA", "SANB11.SA", "SBSP3.SA", "SLCE3.SA", 
        "SMTO3.SA", "SUZB3.SA", "TAEE11.SA", "TIMS3.SA", "TOTS3.SA", "UGPA3.SA", "USIM5.SA", 
        "VALE3.SA", "VBBR3.SA", "BHIA3.SA", "VIVT3.SA", "WEGE3.SA", "YDUQ3.SA"]
quotes = yf.download(assets, start = start, end = end)["Adj Close"]
quotes.head()
quotes.tail()
#drop a column
quotes = quotes.drop(quotes.columns[1], axis=1)
quotes.isna().sum().sum()  # Checking for NAs
pd.set_option('display.max_columns', None)
# Remove '.SA'
quotes.columns = [col[:-3] for col in quotes.columns]
# Assign the number of assets and days
n_days = quotes.shape[0]
n_assets = quotes.shape[1]
# Calculate the number of pairs
num_pairs = math.factorial(n_assets) / (math.factorial(n_assets - 2) * 2)
# Create a DataFrame to the pairs
columns = ["N", "Acao1", "Acao2", "PP1", "PP2", "PPR", "Alfa", "Beta", "pBeta","Preco1", "Preco2", "Max", "Min", "DesvioP", "DesvioAb", "OU"]
pares = pd.DataFrame(columns=columns)
z = 0
p = n_assets

# OU functions - # A shorter half-life implies faster mean reversion
def ornstein_uhlenbeck(params, data):
    theta, mu, sigma = params
    dt = 1.0 / 252  # Assuming daily data, adjust accordingly
    diff = np.diff(data)
    diff_mu = diff - theta * (mu - data[:-1]) * dt
    sum_squared_diff = np.sum(diff_mu**2)
    return sum_squared_diff

def half_life(series):
    initial_params = [0.1, np.mean(series), np.std(series)]
    result = minimize(ornstein_uhlenbeck, initial_params, args=(series,))
    theta, _, _ = result.x
    half_life = np.log(2) / abs(theta)
    return half_life

for i in range(0, p):
    y = quotes.iloc[:, i]
    for j in range(i + 1, p):
        x = quotes.iloc[:, j]
        z = z + 1
        pares.loc[z, "N"] = z
        pares.loc[z, "Acao1"] = quotes.columns[i]
        pares.loc[z, "Acao2"] = quotes.columns[j]
        pares.loc[z, "PP1"] = adfuller(y, autolag='AIC')[1]
        pares.loc[z, "PP2"] = adfuller(x, autolag='AIC')[1]
        x = sm.add_constant(x)
        model = OLS(y, x).fit()
        pares.loc[z, "PPR"] = adfuller(model.resid, autolag='AIC')[1]
        pares.loc[z, "Alfa"] = model.params.iloc[0]
        pares.loc[z, "Beta"] = model.params.iloc[1]
        pares.loc[z, "pBeta"] = model.pvalues.iloc[1]  # Ajuste aqui para acessar a posição desejada
        pares.loc[z, "Preco1"] = quotes.iloc[-1, i]
        pares.loc[z, "Preco2"] = quotes.iloc[-1, j]
        pares.loc[z, "Max"] = max(model.resid)
        pares.loc[z, "Min"] = min(model.resid)
        pares.loc[z, "DesvioP"] = np.std(model.resid)
        pares.loc[z, "DesvioAb"] = quotes.iloc[-1, i] - model.params.iloc[0] - model.params.iloc[1] * quotes.iloc[-1, j]
        pares.loc[z, "OU"] = half_life(model.resid)
        j = j + 1
    i = i + 1

pares

# Filter pairs based on conditions
padf1 = pares[pares['PP1'] > 0.10] #raiz unit da primeira acao
padf2 = padf1[padf1['PP2'] > 0.10] #raiz unit da segunda acao
padf3 = padf2[padf2['PPR'] <= 0.01] #raiz unit dos residuos
padf4 = padf3[abs(padf3['DesvioAb']) > padf3['DesvioP']] #pares abertos
padf5 = padf4[padf4['DesvioAb'] < padf4['Max']] #remove quem esta acima do maximo
padf6 = padf5[padf5['DesvioAb'] > padf5['Min']] #remove quem esta abaixo do minimo
df = padf6[padf6['pBeta'] < 0.01] #remove p-valor do beta superior a 0,10
df['acaoCompra'] = None
df['acaoVende'] = None
df['acaoVende'] = df['Acao1']
df['acaoCompra'] = df['Acao2']
df.loc[df['DesvioAb'] < 0, 'acaoVende'] = df['Acao2']
df.loc[df['DesvioAb'] < 0, 'acaoCompra'] = df['Acao1']

# Agora as contagens
contcompra = df['acaoCompra'].value_counts()
contvenda = df['acaoVende'].value_counts()
compra = contcompra.head(10)
venda = contvenda.head(10)
compradf = pd.DataFrame([compra.index.values, compra.values], index=['Acao', 'Compras']).T
vendadf = pd.DataFrame([venda.index.values, venda.values], index=['Acao', 'Vendas']).T
compratexto = compradf['Acao'].to_csv(index=False, header=False).strip()
vendatexto = vendadf['Acao'].to_csv(index=False, header=False).strip()
print("Número de pares testados:", int(num_pairs))
print("Número de pares cointegrados:", int(padf3.shape[0]))
print("Número de pares cointegrados em abertura:", int(padf4.shape[0]))
print("Número de pares cointegrados em abertura dentro de máximo e mínimo:", int(padf6.shape[0]))
compradf
vendadf

# Finalmente, ordenando por tempo de reversao
df_OU = df.sort_values(by='OU', ascending=True)
df_OU.head(10)
tenpairs = df_OU.head(10)

# Salve o arquivo Excel
tenpairs.to_excel('C:/repo/harpa/tenpairs.xlsx', index=False)
tenpairs.to_excel('C:/repo/quant/tenpairs.xlsx', index=False)




# PCR
import pandas as pd
import requests
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

vencimento = '2025-01-17'
# Lista de feriados no Brasil (pode ser necessário atualizar)
feriados_brasil = [
    '2024-01-01',  # Ano Novo
    '2024-02-12',  # Carnaval
    '2024-02-13',  # Carnaval
    '2024-03-29',  # Paixão de Cristo
    '2024-05-01',  # Dia do Trabalho
    '2024-05-30',  # Corpus Christi
    '2024-11-15',  # Proclamação da República
    '2024-11-20',  # Zumbi
    '2024-12-24',  # Natal
    '2024-12-25',   # Natal
    '2024-12-31',  # Ano novo
]

# Converter feriados para formato de data
feriados_brasil = [datetime.strptime(data, '%Y-%m-%d').date() for data in feriados_brasil]

# Verificar se o dia é útil
def eh_dia_util(data):
    return data.weekday() < 5 and data not in feriados_brasil

# Encontrar o último dia útil
data_ontem = datetime.today().date() - timedelta(days=1)
while not eh_dia_util(data_ontem):
    data_ontem -= timedelta(days=1)
data_ontem


### OS CÓDIGOS PARA CADA AÇÃO

subjacente = 'ABEV3'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrabev.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrabev.xlsx')


subjacente = 'BBDC4'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrbbdc.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrbbdc.xlsx')


subjacente = 'BOVA11'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrbova.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrbova.xlsx')


subjacente = 'PETR4'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrpetr.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrpetr.xlsx')


subjacente = 'VALE3'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
# Um vencimento
      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)
calls = chain[chain['tipo'] == 'CALL']
puts = chain[chain['tipo'] == 'PUT']
soma_callsn = calls['negocios'].sum()
soma_putsn = puts['negocios'].sum()
soma_callsv = calls['volume'].sum()
soma_putsv = puts['volume'].sum()

# Calculate put-call ratio - number of trades
pcr_negocios = soma_putsn / soma_callsn
pcr_volume = soma_putsv / soma_callsv

# Carregar o arquivo Excel existente ou criar um novo
try:
    workbook = load_workbook('C:/repo/harpa/pcrvale.xlsx')
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Data', 'PCRn', 'PCRv'])  # Cabeçalho se for um novo arquivo
# Selecionar a planilha ativa
sheet = workbook.active
# Adicionar as informações capturadas à planilha
sheet.append([data_ontem, pcr_negocios, pcr_volume])
# Salvar as alterações no arquivo Excel
workbook.save('C:/repo/harpa/pcrvale.xlsx')






# Seguro da carteira

import pandas as pd
import requests
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# Option chain de BOVA11
subjacente = 'BOVA11'   # não vem do Yfinance! esqueça o ".SA" ao final do ticker!
vencimento = '2025-01-17'      # YYYY-MM-DD
def optionchaindate(subjacente, vencimento):
    url = f'https://opcoes.net.br/listaopcoes/completa?idAcao={subjacente}&listarVencimentos=false&cotacoes=true&vencimentos={vencimento}'
    r = requests.get(url).json()
    x = ([subjacente, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8], i[9], i[10]] for i in r['data']['cotacoesOpcoes'])
    return pd.DataFrame(x, columns=['subjacente', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco', 'negocios', 'volume'])

chain = optionchaindate(subjacente, vencimento)

# Filtros
symbol = 'BOVA11.SA'
stock_data = yf.download(symbol)['Adj Close']
last_price = stock_data.iloc[-1]
chain2 = chain[chain['tipo'] == 'PUT']
chain3 = chain2[chain2['strike'] < last_price*0.85]
chain4 = chain3.sort_values(by='strike', ascending=False)
chain5 = chain4.drop('modelo', axis=1)
chain6 = chain5.reset_index(drop=True)
bova11 = chain6.head(10)

# Renomeie as colunas
novo_nome_colunas = {
    'subjacente': 'Subjacente',
    'vencimento': 'Vencimento',
    'ativo': 'Ativo',
    'tipo': 'Tipo',
    'strike': 'Strike',
    'preco': 'Preço',
    'negocios': 'Negócios',
    'volume': 'Volume'
}

bova11 = bova11.rename(columns=novo_nome_colunas)

wb = Workbook()
ws = wb.active

# Adicione os dados do DataFrame ao arquivo Excel
for row in dataframe_to_rows(bova11, index=False, header=True):
    ws.append(row)

# Salve o arquivo Excel
wb.save('C:/repo/harpa/bova11_disaster.xlsx')

wb.save('C:/repo/quant/bova11_disaster.xlsx')
